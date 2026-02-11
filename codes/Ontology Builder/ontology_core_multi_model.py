# ontology_core_multi_model.py
# ------------------------------------------------------------
# Multi-input pipeline logic for PDF / TTL / OWL / Excel -> LLM build
# with provider selection: OpenAI / Anthropic / Gemini
# ------------------------------------------------------------

import os
import re
import json
from typing import List, Dict, Set, Tuple, Optional
from datetime import date
from collections import Counter

import fitz  # PyMuPDF
from openai import OpenAI
from openpyxl import load_workbook

from rdflib import Graph, RDF, RDFS, OWL, Namespace, URIRef
from rdflib.namespace import SKOS
import requests
from rdflib.term import Literal

from rdflib.namespace import DC, DCTERMS
SCHEMA = Namespace("http://schema.org/")

# ---------------------------
# helpers
# ---------------------------
def _strip_code_fences(s: str) -> str:
    s = (s or "").strip()
    if s.startswith("```"):
        lines = s.splitlines()
        # handles ```turtle as well
        if lines and lines[-1].strip() == "```":
            return "\n".join(lines[1:-1]).strip()
        # fallback: if starts with ``` and later contains closing fence
        for i in range(1, len(lines)):
            if lines[i].strip() == "```":
                return "\n".join(lines[1:i]).strip()
    return s


# ---------------------------
# Provider calls
# ---------------------------
def _call_openai(api_key: str, model: str, system_text: str, user_prompt: str) -> str:
    client = OpenAI(api_key=api_key)
    resp = client.responses.create(
        model=model,
        input=[
            {"role": "system", "content": system_text},
            {"role": "user", "content": user_prompt},
        ],
        reasoning={"effort": "high"},
        text={"verbosity": "low"},
    )
    return resp.output_text


def _call_anthropic(api_key: str, model: str, system_text: str, user_prompt: str) -> str:
    """
    Anthropic Messages API (direct HTTP). Requires: pip install requests
    Uses the same system/user content; prompt content is unchanged.
    """
    url = "https://api.anthropic.com/v1/messages"
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    payload = {
        "model": model,
        "max_tokens": 8000,
        "system": system_text,
        "messages": [{"role": "user", "content": user_prompt}],
    }
    r = requests.post(url, headers=headers, json=payload, timeout=300)
    if r.status_code >= 300:
        raise RuntimeError(f"Anthropic API error {r.status_code}: {r.text}")

    data = r.json()
    parts = []
    for blk in data.get("content", []):
        if blk.get("type") == "text":
            parts.append(blk.get("text", ""))
    return "".join(parts).strip()


def _call_gemini(api_key: str, model: str, system_text: str, user_prompt: str) -> str:
    """
    Gemini Generative Language API (direct HTTP).
    Model examples: gemini-1.5-pro, gemini-1.5-flash
    """
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
    payload = {
        "contents": [
            {
                "role": "user",
                "parts": [{"text": f"{system_text}\n\n{user_prompt}"}],
            }
        ],
        "generationConfig": {
            "temperature": 0.2,
            "maxOutputTokens": 8192,
        },
    }
    r = requests.post(url, json=payload, timeout=300)
    if r.status_code >= 300:
        raise RuntimeError(f"Gemini API error {r.status_code}: {r.text}")

    data = r.json()
    cands = data.get("candidates", [])
    if not cands:
        raise RuntimeError(f"Gemini returned no candidates: {data}")

    parts = cands[0].get("content", {}).get("parts", [])
    txt = "".join(p.get("text", "") for p in parts)
    return txt.strip()


def _call_llm(provider: str, api_key: str, model: str, system_text: str, user_prompt: str) -> str:
    provider_norm = (provider or "").strip().lower()
    if provider_norm in ("openai", "gpt", "chatgpt"):
        return _call_openai(api_key, model, system_text, user_prompt)
    if provider_norm in ("anthropic", "claude"):
        return _call_anthropic(api_key, model, system_text, user_prompt)
    if provider_norm in ("gemini", "google"):
        return _call_gemini(api_key, model, system_text, user_prompt)
    raise ValueError(f"Unknown provider '{provider}'. Use OpenAI, Anthropic, or Gemini.")


def _clean_ws(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\u00A0", " ")
    s = re.sub(r"(\w)-\n(\w)", r"\1\2", s)
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r" +\n", "\n", s)
    s = re.sub(r"\n{2,}", "\n\n", s)
    return s.strip()

_DEF_NAME_RX = re.compile(r"(definition|description|textdefinition|gloss|documentation|explanation|note)", re.IGNORECASE)

def _local_name(uri: str) -> str:
    # fragment after # or last path segment
    if "#" in uri:
        return uri.rsplit("#", 1)[-1]
    return uri.rsplit("/", 1)[-1]

def discover_definition_predicates(g: Graph) -> List[URIRef]:
    """
    Ordered predicates:
      - skos:definition (+ its subproperties)
      - IAO:0000115 (+ its subproperties)
      - dcterms:description / dc:description / schema:description (+ subproperties)
      - heuristic definition-like annotation properties
      - rdfs:comment (always last)
    """
    bases = [
        SKOS.definition,
        OBO["IAO_0000115"],
        DCTERMS.description,
        DC.description,
        SCHEMA.description,
    ]

    out: List[URIRef] = []
    seen: Set[URIRef] = set()

    def add(p: URIRef):
        if p not in seen:
            seen.add(p)
            out.append(p)

    # helper: collect subproperties of a given property (transitive)
    def subprops_of(base: URIRef) -> List[URIRef]:
        acc: Set[URIRef] = set()
        frontier = [base]
        while frontier:
            cur = frontier.pop()
            for child in g.subjects(RDFS.subPropertyOf, cur):
                if isinstance(child, URIRef) and child not in acc:
                    acc.add(child)
                    frontier.append(child)
        # deterministic order: sort by URI string
        return sorted(acc, key=lambda u: str(u))

    # 1) bases, each followed immediately by its subproperties
    for b in bases:
        add(b)
        for sp in subprops_of(b):
            add(sp)


    # 2) heuristic “definition-like” annotation properties (only if not already included)
    candidates = set(g.subjects(RDF.type, OWL.AnnotationProperty))
    candidates |= set(g.subjects(RDF.type, RDF.Property))

    # also include predicates that are USED with Literal objects (even if untyped)
    for s, p, o in g.triples((None, None, None)):
        if isinstance(p, URIRef) and isinstance(o, Literal):
            candidates.add(p)

    def looks_definition_like(p: URIRef) -> bool:
        p_str = str(p)
        name = _local_name(p_str)
        labels = [str(lit) for lit in g.objects(p, RDFS.label)]
        return _DEF_NAME_RX.search(name) or any(_DEF_NAME_RX.search(l) for l in labels)

    for p in sorted([p for p in candidates if isinstance(p, URIRef)], key=lambda u: str(u)):
        if p not in seen and looks_definition_like(p):
            add(p)

    # 3) final fallback
    add(RDFS.comment)

    return out


# ----------------------------------------------------------------------
# Output hardening 
# ----------------------------------------------------------------------
_CANON_PREFIXES = {
    "rdf": "http://www.w3.org/1999/02/22-rdf-syntax-ns#",
    "rdfs": "http://www.w3.org/2000/01/rdf-schema#",
    "owl": "http://www.w3.org/2002/07/owl#",
    "xsd": "http://www.w3.org/2001/XMLSchema#",
    "skos": "http://www.w3.org/2004/02/skos/core#",
    "dc": "http://purl.org/dc/elements/1.1/",
    "dcterms": "http://purl.org/dc/terms/",
    # we keep dct as alias because some models output dct:
    "dct": "http://purl.org/dc/terms/",
}


def _normalize_standard_prefixes(ttl: str) -> str:
    """
    Fix common model failure mode: wrong namespace URI for standard prefixes (esp. owl:).
    This does NOT change your prompt; it repairs the output.
    """
    lines = (ttl or "").splitlines()
    out = []
    prefix_rx = re.compile(r"^\s*@prefix\s+([A-Za-z][\w-]*)\s*:\s*<([^>]+)>\s*\.\s*$")

    seen = set()
    for ln in lines:
        m = prefix_rx.match(ln)
        if not m:
            out.append(ln)
            continue

        pfx = m.group(1)
        uri = m.group(2)

        # enforce canonical URI for known prefixes
        if pfx in _CANON_PREFIXES:
            canon = _CANON_PREFIXES[pfx]
            if uri != canon:
                ln = f"@prefix {pfx}: <{canon}> ."
        seen.add(pfx)
        out.append(ln)

    # ensure required standard prefixes exist (only add if missing)
    # Avoids rewriting user-specific prefixes.
    insert_at = 0
    while insert_at < len(out) and out[insert_at].strip().startswith("@prefix"):
        insert_at += 1

    missing_lines = []
    for pfx, uri in _CANON_PREFIXES.items():
        if pfx not in seen:
            missing_lines.append(f"@prefix {pfx}: <{uri}> .")

    if missing_lines:
        out = out[:insert_at] + missing_lines + out[insert_at:]

    return "\n".join(out).strip() + "\n"


def _validate_bfo_import(ttl_path: str, ontology_iri: str) -> None:
    """
    Semantic validation: ensures the output contains a real OWL import triple.
    This catches the Claude-style 'owl:' prefix misbinding immediately.
    """
    bfo_url = URIRef("http://purl.obolibrary.org/obo/bfo/2020/bfo.owl")
    ont = URIRef(ontology_iri)

    g = Graph()
    g.parse(ttl_path, format="turtle")

    if (ont, OWL.imports, bfo_url) not in g:
        raise ValueError(
            "Generated TTL does not contain the required OWL import triple:\n"
            f"  <{ontology_iri}> owl:imports <http://purl.obolibrary.org/obo/bfo/2020/bfo.owl>\n"
            "This usually happens when the model binds 'owl:' to the wrong namespace."
        )


# ----------------------------------------------------------------------
# PDF extraction (Section 3 terms and definitions)
# ----------------------------------------------------------------------
def _rebuild_text_blocks(pdf_path: str) -> str:
    pages = []
    with fitz.open(pdf_path) as d2:
        for pno in range(len(d2)):
            page = d2[pno]
            blocks = page.get_text("blocks")
            blocks_sorted = sorted(blocks, key=lambda b: (round(b[1], 1), round(b[0], 1)))
            text = "\n".join(b[4] for b in blocks_sorted if str(b[4]).strip())
            pages.append(text)
    return "\n\n".join(pages)


def _looks_like_toc(ln: str) -> bool:
    s = ln.strip()
    if re.search(r"\.{5,}\s*\d+$", s):
        return True
    if re.search(r"\.{10,}", s):
        return True
    return False


def isolate_section3(text: str) -> str:
    t = text.replace("\u00A0", " ")
    lines = t.splitlines()

    offsets, pos = [], 0
    for i, ln in enumerate(lines):
        offsets.append((pos, pos + len(ln), i))
        pos += len(ln) + 1

    candidate_spans = []
    for i, ln in enumerate(lines):
        if re.search(r"(?i)^\s*3\s*$", ln):
            if i + 1 < len(lines) and re.search(r"(?i)^\s*terms?\s+and\s+definitions\b", lines[i + 1]):
                if not (_looks_like_toc(lines[i]) or _looks_like_toc(lines[i + 1])):
                    candidate_spans.append((offsets[i][0], offsets[i + 1][1]))
        if re.search(r"(?i)^\s*3[\s.\u00A0]+terms?\s+and\s+definitions\b", ln):
            if not _looks_like_toc(ln):
                candidate_spans.append((offsets[i][0], offsets[i][1]))

    def _next_term_heading_pos(start_char: int) -> int:
        m = re.search(r"(?m)^\s*3\s*\.\s*\d", t[start_char:start_char + 4000])
        return (start_char + m.start()) if m else -1

    best_start, best_follow = None, None
    for s0, s1 in candidate_spans:
        pos3x = _next_term_heading_pos(s1)
        if pos3x != -1 and (best_follow is None or pos3x < best_follow):
            best_start, best_follow = s0, pos3x

    if best_start is None:
        m_any = re.search(r"(?m)^\s*3\s*\.\s*\d", t)
        if not m_any:
            return ""
        best_start = m_any.start()

    m_end = None
    for j, ln in enumerate(lines):
        s, e, _ = offsets[j]
        if e <= best_start:
            continue
        if re.search(r"(?m)^\s*[4-9][\s.\u00A0]+[^\n]+$", ln):
            if _looks_like_toc(ln):
                continue
            m_end = s
            break

    end_char = m_end if m_end is not None else len(t)
    return t[best_start:end_char].strip()


def extract_terms_with_definitions(text: str) -> List[Dict[str, str]]:
    if not text:
        return []

    text = _clean_ws(text)

    heading_rx = re.compile(r"(?m)^\s*3(?:[.\u00A0]\s*\d+)+(?:\s+.*)?$")
    num_rx = re.compile(r"^\s*(3(?:[.\u00A0]\s*\d+)+)\b\s*(.*)$")
    note_rx = re.compile(r"(?im)^\s*note\s+\d*\s+to\s+entry:.*$")
    _SPACED_DASH_RX = re.compile(r"\s[–—-]\s")

    def split_label_inline(s: str) -> Tuple[str, str]:
        s = s.strip()
        if ":" in s:
            i = s.index(":")
            return s[:i].strip(), s[i + 1:].strip()
        m = _SPACED_DASH_RX.search(s)
        if m:
            left = s[:m.start()].strip()
            right = s[m.end():].strip()
            return left, right
        return s, ""

    items: List[Dict[str, str]] = []
    matches = list(heading_rx.finditer(text))
    for i, m in enumerate(matches):
        start = m.end()
        end = matches[i + 1].start() if (i + 1) < len(matches) else len(text)

        line = m.group(0)
        mnum = num_rx.match(line)
        if not mnum:
            continue

        item_num = mnum.group(1).replace("\u00A0", "").replace(" ", "")
        tail = mnum.group(2).strip()
        block = text[start:end].strip()

        if len(tail) < 2:
            for ln in block.splitlines():
                if ln.strip() and not ln.strip().startswith("3."):
                    tail = ln.strip()
                    break

        label, inline_def = split_label_inline(tail)
        label = re.sub(r"\s*\([^()]*\)\s*$", "", label).strip()

        lines = block.splitlines()
        lines_wo_notes = [ln for ln in lines if not note_rx.match(ln)]
        para = []
        for ln in lines_wo_notes:
            s = ln.strip()
            if not s and para:
                break
            if s.startswith("3."):
                break
            if s.lower().startswith(("example", "examples")):
                break
            para.append(s)
        block_para = " ".join([ln for ln in para if ln]).strip()

        parts = []
        if inline_def:
            parts.append(inline_def)
        if block_para and (not inline_def or block_para not in inline_def):
            parts.append(block_para)

        definition = " ".join(parts)

        for pat in [
            r"©\s*\S+\s*\d{4}.*?(?:reserved|rights reserved).*",
            r"INTERNATIONAL\s+STANDARD\b.*",
            r"This\s+preview\s+is\s+downloaded.*",
            r"https?://\S+",
            r"\bAll\s+rights\s+reserved\b.*",
        ]:
            definition = re.sub(pat, "", definition, flags=re.IGNORECASE)

        definition = _clean_ws(definition)

        if 2 <= len(label) <= 200:
            items.append({"num": item_num, "label": label, "definition": definition})

    out, seen = [], set()
    for it in items:
        k = it["label"].lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(it)
    return out


# ----------------------------------------------------------------------
# Excel loader (.xlsx/.xls) with Term/Definition columns
# ----------------------------------------------------------------------
def load_excel_terms(xlsx_path: str) -> List[Dict[str, str]]:
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("Excel file has no header row.")

    header = [(_clean_ws(cell) if cell is not None else "") for cell in header_row]
    header_lc = [h.lower() for h in header]

    def _find_col(name: str) -> Optional[int]:
        name_lc = name.lower()
        for i, h in enumerate(header_lc):
            if h == name_lc:
                return i
        return None

    term_col = _find_col("Term")
    if term_col is None:
        raise ValueError("Excel must contain a header column named 'Term'.")

    def_col = _find_col("Definition")
    terms: List[Dict[str, str]] = []
    idx = 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        term_val = row[term_col] if term_col < len(row) else None
        def_val = row[def_col] if (def_col is not None and def_col < len(row)) else None

        if term_val is None:
            continue

        label = _clean_ws(term_val)
        if not label:
            continue

        definition = _clean_ws(def_val) if def_val is not None else ""
        terms.append({"num": str(idx), "label": label, "definition": definition})
        idx += 1

    if not terms:
        raise ValueError("No terms found in Excel. Ensure rows under 'Term' are not empty.")

    return terms


# ----------------------------------------------------------------------
# TTL/OWL RDF class extraction
# ----------------------------------------------------------------------
OBO = Namespace("http://purl.obolibrary.org/obo/")
VANN = Namespace("http://purl.org/vocab/vann/")

OBO_PURL_DIR = "http://purl.obolibrary.org/obo/"
OBOINOWL = Namespace("http://www.geneontology.org/formats/oboInOwl#")

def _local_name_from_uri(uri: str) -> str:
    # fragment after # or last path segment
    if "#" in uri:
        return uri.rsplit("#", 1)[-1]
    return uri.rsplit("/", 1)[-1]

def _infer_obo_idspace_from_ontology_iri(ont_iri: str) -> Optional[str]:
    """
    If ontology IRI is like .../obo/label.owl
    """
    if not ont_iri:
        return None
    s = ont_iri.split("#", 1)[0]
    m = re.match(r"^http://purl\.obolibrary\.org/obo/([^/]+)\.(owl|ttl|rdf|xml|obo)$", s, flags=re.IGNORECASE)
    if not m:
        return None
    return m.group(1).upper().strip()

def _infer_obo_idspace_from_graph(g: Graph) -> Optional[str]:
    """
    Prefer explicit oboInOwl:idspace, else fall back to ontology IRI filename.
    """
    # 1) explicit idspace
    for ont in g.subjects(RDF.type, OWL.Ontology):
        for v in g.objects(ont, OBOINOWL.idspace):
            s = str(v).strip()
            if s:
                return s.upper()

    # 2) from ontology IRI
    for ont in g.subjects(RDF.type, OWL.Ontology):
        cand = _infer_obo_idspace_from_ontology_iri(str(ont))
        if cand:
            return cand

    return None

def _guess_namespace(uri: str) -> str:
    i_hash = uri.rfind("#")
    i_slash = uri.rfind("/")
    i = max(i_hash, i_slash)
    if i == -1:
        return uri
    return uri[: i + 1]


def _ontology_dir_namespace(ont_iri: str) -> Optional[str]:
    if not ont_iri:
        return None
    ont_iri = ont_iri.split("#", 1)[0]

    if re.search(r"\.(owl|ttl|rdf|xml|n3|nt|trig|obo)$", ont_iri, flags=re.IGNORECASE):
        if "/" in ont_iri:
            dir_ns = ont_iri.rsplit("/", 1)[0] + "/"

            # never use the shared OBO directory as "local"
            if dir_ns == OBO_PURL_DIR:
                return None

            return dir_ns
    return None


def _get_local_namespace_uris(graph: Graph) -> Set[str]:
    local_ns_uris: Set[str] = set()

    for ont in graph.subjects(RDF.type, OWL.Ontology):
        for ns in graph.objects(ont, VANN.preferredNamespaceUri):
            local_ns_uris.add(str(ns))

    if not local_ns_uris:
        prefixes = set()
        for ont in graph.subjects(RDF.type, OWL.Ontology):
            for lit in graph.objects(ont, VANN.preferredNamespacePrefix):
                prefixes.add(str(lit))

        if prefixes:
            for prefix, ns in graph.namespace_manager.namespaces():
                if prefix in prefixes:
                    local_ns_uris.add(str(ns))

    if not local_ns_uris:
        for ont in graph.subjects(RDF.type, OWL.Ontology):
            ont_iri = str(ont)
            if not ont_iri:
                continue

            stripped = ont_iri.rstrip("/#")
            if stripped:
                local_ns_uris.update({stripped, stripped + "/", stripped + "#"})

            # also consider the ontology directory namespace
            dir_ns = _ontology_dir_namespace(ont_iri)
            if dir_ns:
                local_ns_uris.add(dir_ns)
                local_ns_uris.add(dir_ns + "#")

    if not local_ns_uris:
        ns_counts = Counter()
        for cls in set(graph.subjects(RDF.type, OWL.Class)):
            if cls.__class__.__name__ == "BNode":
                continue
            uri = str(cls)
            ns = _guess_namespace(uri)
            ns_counts[ns] += 1

        if ns_counts:
            max_count = max(ns_counts.values())
            for ns, count in ns_counts.items():
                if count >= max_count * 0.5 and count >= 1:
                    local_ns_uris.add(ns)

    if not local_ns_uris:
        for prefix, ns in graph.namespace_manager.namespaces():
            if prefix == "":
                local_ns_uris.add(str(ns))

    return local_ns_uris

def _axiom_annotation_values(g: Graph, cls: URIRef, prop: URIRef) -> List[str]:
    vals: List[str] = []
    for ax in g.subjects(RDF.type, OWL.Axiom):
        if (ax, OWL.annotatedSource, cls) in g and (ax, OWL.annotatedProperty, prop) in g:
            for tgt in g.objects(ax, OWL.annotatedTarget):
                vals.append(str(tgt))
    return vals

def _collect_class_info(graph: Graph, subject, def_props: List[URIRef]):
    labels = [str(o) for o in graph.objects(subject, RDFS.label)]
    if labels:
        label = labels[0]
    else:
        try:
            label = subject.n3(graph.namespace_manager)
        except Exception:
            label = str(subject)

    definitions: List[str] = []
    for prop in def_props:
        # direct annotation values
        for o in graph.objects(subject, prop):
            definitions.append(str(o))

        # axiom-reified annotation values
        if isinstance(subject, URIRef):
            definitions.extend(_axiom_annotation_values(graph, subject, prop))

    return label, definitions


def _read_xml_base_if_present(path: str) -> Optional[str]:
    try:
        with open(path, "rb") as f:
            head = f.read(20000)
        txt = head.decode("utf-8", errors="ignore")
        m = re.search(r'xml:base\s*=\s*"([^"]+)"', txt)
        if m:
            return m.group(1).strip()
    except Exception:
        pass
    return None

def _augment_namespaces_for_obo(local_ns_uris: Set[str]) -> Set[str]:
    # Do not automatically add the global /obo/ namespace
    return set(local_ns_uris)

def _infer_dominant_id_prefix(class_uris: List[str]) -> Optional[str]:
    rx = re.compile(r"^([A-Za-z][A-Za-z0-9]*(?:_[A-Za-z0-9]+)*)_\d+$")
    prefixes = []
    for uri in class_uris:
        local = uri.rsplit("/", 1)[-1]
        m = rx.match(local)
        if m:
            prefixes.append(m.group(1))
    if not prefixes:
        return None
    c = Counter(prefixes)
    top, top_count = c.most_common(1)[0]
    return top if top_count >= 10 else None


def _looks_like_html_file(path: str) -> bool:
    try:
        with open(path, "rb") as f:
            head = f.read(4096)
        s = head.decode("utf-8", errors="ignore").lower()
        return ("<html" in s) or ("<!doctype html" in s) or ("<style" in s)
    except Exception:
        return False


def _parse_rdf_graph(path: str) -> Tuple[Graph, str]:
    # Catch "TTL" files that are actually HTML downloads
    if _looks_like_html_file(path):
        raise ValueError(
            f"File '{os.path.basename(path)}' looks like HTML, not Turtle/RDF. "
            "This often happens when you downloaded a web page instead of the raw .ttl. "
            "Please re-download the raw TTL content."
        )

    ext = os.path.splitext(path)[1].lower()

    if ext == ".ttl":
        fmts = ["turtle", "n3"]
    elif ext == ".owl":
        fmts = ["xml", "application/rdf+xml", "turtle", "n3"]
    else:
        fmts = ["xml", "turtle", "n3", "nt", "trig"]

    last_err = None
    for fmt in fmts:
        g = Graph()
        try:
            g.parse(path, format=fmt)
            return g, fmt
        except Exception as e:
            last_err = e

    g = Graph()
    try:
        g.parse(path)
        return g, "auto"
    except Exception as e:
        raise ValueError(
            f"Failed to parse RDF file '{path}'. Last error: {last_err}. Auto error: {e}"
        ) from e


def load_rdf_terms_with_hierarchy(rdf_path: str) -> Tuple[List[Dict[str, str]], Dict[str, str]]:
    """
    Loads local owl:Class terms and extracts local subclass hierarchy.
    Returns:
      - terms: List[{num,label,definition}]
      - parent_map: child_num -> parent_num ("" if none)
    Only considers rdfs:subClassOf edges where both endpoints are local classes.
    """
    g, used_fmt = _parse_rdf_graph(rdf_path)
    def_props = discover_definition_predicates(g)

    class_nodes = []
    class_uris = []
    for cls in set(g.subjects(RDF.type, OWL.Class)):
        if cls.__class__.__name__ == "BNode":
            continue
        class_nodes.append(cls)
        class_uris.append(str(cls))

    if not class_nodes:
        raise ValueError(
            f"No owl:Class found in RDF ({os.path.basename(rdf_path)}). Parsed format={used_fmt}."
        )

    local_ns_uris = _get_local_namespace_uris(g)

    if os.path.splitext(rdf_path)[1].lower() == ".owl":
        xml_base = _read_xml_base_if_present(rdf_path)
        if xml_base:
            stripped = xml_base.rstrip("/#")
            local_ns_uris.update({stripped, stripped + "/", stripped + "#"})

    local_ns_uris = _augment_namespaces_for_obo(local_ns_uris)
    dominant_prefix = _infer_dominant_id_prefix(class_uris)

    def _select_local_classes() -> List:
        # if this is an OBO ontology, local classes are those in its IDSPACE.
        idspace = _infer_obo_idspace_from_graph(g)
        if idspace:
            idspace_selected = []
            for cls in class_nodes:
                u = str(cls)
                if _local_name_from_uri(u).startswith(idspace + "_"):
                    idspace_selected.append(cls)
            if idspace_selected:
                return idspace_selected

        selected = []

        # 1) namespace-based selection (non-OBO case)
        if local_ns_uris:
            for cls in class_nodes:
                u = str(cls)
                if any(u.startswith(ns) for ns in local_ns_uris):
                    selected.append(cls)

        # If namespace selection succeeded, KEEP it
        if selected:
            return selected

        # If namespace selection finds nothing, start from all classes
        selected = class_nodes[:]  # fallback

        # Dominant prefix is a fallback ONLY (used when namespace selection failed)
        if dominant_prefix:
            pref_selected_all = []
            for cls in class_nodes:
                local = _local_name_from_uri(str(cls))
                if local.startswith(dominant_prefix + "_"):
                    pref_selected_all.append(cls)

            if len(pref_selected_all) >= 10:
                return pref_selected_all

        return selected


    selected_nodes = _select_local_classes()
    selected_nodes = sorted(selected_nodes, key=lambda x: str(x))  # deterministic

    # Build terms + uri<->num maps
    terms: List[Dict[str, str]] = []
    uri_to_num: Dict[str, str] = {}
    num_to_uri: Dict[str, str] = {}

    for idx, cls in enumerate(selected_nodes, start=1):

        label, defs = _collect_class_info(g, cls, def_props)
        # clean and drop empty defs
        defs = [_clean_ws(d) for d in defs]
        defs = [d for d in defs if d]

        selected_def = ""
        if defs:
            for d in defs:
                if not d.strip().lower().startswith("aristotelian:"):
                    selected_def = d
                    break
            if not selected_def:
                selected_def = defs[0]

        num = str(idx)
        uri = str(cls)
        uri_to_num[uri] = num
        num_to_uri[num] = uri
        terms.append({"num": num, "label": _clean_ws(label), "definition": _clean_ws(selected_def)})

    if not terms:
        raise ValueError(
            f"No classes extracted from RDF ({os.path.basename(rdf_path)}). Parsed format={used_fmt}."
        )

    # Extract local subclass relations: child -> one local parent (if any)
    parent_map: Dict[str, str] = {t["num"]: "" for t in terms}
    for cls in selected_nodes:
        child_uri = str(cls)
        child_num = uri_to_num.get(child_uri)
        if not child_num:
            continue

        local_parents: List[str] = []
        for p in g.objects(cls, RDFS.subClassOf):
            if p.__class__.__name__ == "BNode":
                continue
            p_uri = str(p)
            if p_uri in uri_to_num:
                local_parents.append(p_uri)

        if local_parents:
            chosen_parent_uri = sorted(set(local_parents))[0]
            parent_map[child_num] = uri_to_num[chosen_parent_uri]

    return terms, parent_map


# ----------------------------------------------------------------------
# Metadata fallback
# ----------------------------------------------------------------------
def derive_ontology_metadata_from_input(input_path: str) -> Dict[str, str]:
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    domain = base_name or "input source"
    title = f"{domain} ontology"
    label_abbrev = "".join(w[0] for w in title.split() if w and w[0].isalpha()).lower() or "onto"
    ontology_iri = f"https://github.com/ISE-FIZKarlsruhe/LLM4BFO/{label_abbrev}"
    desc = (
        "Ontology created from an input file containing classes and definitions. "
        "It encodes domain classes and maps each class to BFO 2020."
    )
    desc = " ".join(desc.split())
    version = date.today().strftime("%Y.%m.%d") + ".LLM"
    return {
        "title": title,
        "label": label_abbrev,
        "description": desc,
        "version": version,
        "domain": domain,
        "ontology_iri": ontology_iri,
    }


# ----------------------------------------------------------------------
# main callable for GUI
# ----------------------------------------------------------------------
def run_pipeline(
    input_path: str,
    ttl_out: str,
    provider: str,
    api_key: str,
    model_name: str,
    meta: Dict[str, str] | None,
) -> str:
    logs: List[str] = []

    def log(msg: str):
        logs.append(msg)
        print(msg)

    try:
        if not api_key:
            raise RuntimeError("Please provide API key.")
        if not input_path or not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")

        ext = os.path.splitext(input_path)[1].lower()
        log(f"Input file: {input_path} (ext={ext})")

        model_label = model_name
        log(f"Loaded API key, provider: {provider}, model: {model_label}")

        META = meta if meta is not None else derive_ontology_metadata_from_input(input_path)

        # ----------------
        # Load terms
        # ----------------
        if ext == ".pdf":
            with fitz.open(input_path) as doc:
                full_text = "\n".join(page.get_text() for page in doc)

            terms_section = isolate_section3(full_text)
            if not terms_section.strip():
                rebuilt_text = _rebuild_text_blocks(input_path)
                terms_section = isolate_section3(rebuilt_text)

            terms = extract_terms_with_definitions(terms_section)
            if not terms:
                raise ValueError("No terms extracted from PDF Section 3.")

            def _parent_num(n: str) -> str:
                parts = n.split(".")
                return ".".join(parts[:-1]) if len(parts) > 2 else ""

            parent_map: Dict[str, str] = {}
            nums = {t["num"] for t in terms}
            for t in terms:
                p = _parent_num(t["num"])
                parent_map[t["num"]] = p if p in nums else ""

            log(f"Extracted {len(terms)} term(s) from PDF Section 3:")
            for t in terms:
                log(f" - {t['num']}: {t['label']}")

        elif ext in (".xlsx", ".xls"):
            terms = load_excel_terms(input_path)
            parent_map = {t["num"]: "" for t in terms}

            log(f"Loaded {len(terms)} term(s) from Excel:")
            for t in terms:
                log(f" - {t['num']}: {t['label']}")

        elif ext in (".ttl", ".owl"):
            terms, parent_map = load_rdf_terms_with_hierarchy(input_path)

            log(f"Loaded {len(terms)} term(s) from RDF (TTL/OWL):")
            for t in terms:
                log(f" - {t['num']}: {t['label']}")

        else:
            raise ValueError(f"Unsupported input file type: {ext}. Use .pdf, .ttl, .owl, .xlsx, or .xls.")

        # ----------------
        # BFO reference text 
        # ----------------
        BFO_TEXT = r"""
Here are the definitions of categories of the bfo upper ontology:

- entity (http://purl.obolibrary.org/obo/BFO_0000001)
  elucidation: an entity is anything that exists or has existed or will exist

- continuant (http://purl.obolibrary.org/obo/BFO_0000002)
  elucidation: a continuant is an entity that persists, endures, or continues to exist through time while maintaining its identity

- generically dependent continuant (http://purl.obolibrary.org/obo/BFO_0000031)
  elucidation: a generically dependent continuant is an entity that exists in virtue of the fact that there is at least one of what may be multiple copies; it is the content or the pattern that the multiple copies share

- independent continuant (http://purl.obolibrary.org/obo/BFO_0000004)
  definition: an independent continuant is a continuant which is such that there is no c such that it specifically depends on c and no c such that it generically depends on c

- immaterial entity (http://purl.obolibrary.org/obo/BFO_0000141)
  definition: an immaterial entity is an independent continuant which is such that there is no time t when it has a material entity as continuant part at t

- continuant fiat boundary (http://purl.obolibrary.org/obo/BFO_0000140)
  elucidation: an immaterial entity of zero, one or two dimensions, whose location is determined in relation to some material entity

- fiat line (http://purl.obolibrary.org/obo/BFO_0000142)
  elucidation: a one-dimensional continuant fiat boundary that is continuous

- fiat surface (http://purl.obolibrary.org/obo/BFO_0000146)
  elucidation: a two-dimensional continuant fiat boundary that is self-connected

- fiat point (http://purl.obolibrary.org/obo/BFO_0000147)
  elucidation: a zero-dimensional continuant fiat boundary that consists of a single point

- site (http://purl.obolibrary.org/obo/BFO_0000029)
  elucidation: a three-dimensional immaterial entity whose boundaries are determined relative to some material entity

- spatial region (http://purl.obolibrary.org/obo/BFO_0000006)
  elucidation: a continuant entity that is a continuant part of the spatial projection of spacetime at a given time

- one-dimensional spatial region (http://purl.obolibrary.org/obo/BFO_0000026)
- two-dimensional spatial region (http://purl.obolibrary.org/obo/BFO_0000009)
- three-dimensional spatial region (http://purl.obolibrary.org/obo/BFO_0000028)
- zero-dimensional spatial region (http://purl.obolibrary.org/obo/BFO_0000018)

- material entity (http://purl.obolibrary.org/obo/BFO_0000040)
  elucidation: an independent continuant that always has some portion of matter as continuant part

- fiat object part (http://purl.obolibrary.org/obo/BFO_0000024)

- object (http://purl.obolibrary.org/obo/BFO_0000030)
  elucidation: an object is a material entity that is maximally self-connected, countable, and wholly present at any time it exists. It bears qualities, participates in processes, and does not have temporal parts. Typical denotations are artifacts, devices, tools, containers, fixtures, instruments, components, and samples.

- object aggregate (http://purl.obolibrary.org/obo/BFO_0000027)

- specifically dependent continuant (http://purl.obolibrary.org/obo/BFO_0000020)
  definition: a continuant that inheres in or depends on some independent continuant

- quality (http://purl.obolibrary.org/obo/BFO_0000019)
  elucidation: a quality is a specifically dependent continuant that is continuously present in its bearer and does not require any triggering process for its existence. 
  qualities include directly observable features and setup parameters such as rate, angle, or temperature. 
  examples: mass, length, color, density, rate.
  rule: if the property exists whenever the bearer (or test setup) exists, classify as quality.
  Qualities also include physical states, conditions, or directly observable features such as color, consistency, state, texture, shape, temperature, hardness and softness.
  A quality is not an independent object but inheres in some bearer.

- relational quality (http://purl.obolibrary.org/obo/BFO_0000145)
  definition: a relational quality is a quality that depends on at least two distinct bearers.
  example: distance between two points.

- realizable entity (http://purl.obolibrary.org/obo/BFO_0000017)
  elucidation: a realizable entity is a specifically dependent continuant that is only manifested under certain conditions or in the course of a process. 
  rule: if the property expresses a potential, capacity, or is only measurable during a process/test, it is a realizable entity.

- disposition (http://purl.obolibrary.org/obo/BFO_0000016)
  elucidation: a disposition is a realizable entity such that if it ceases to exist, the bearer is physically changed. its realization occurs when the bearer is in certain physical circumstances, due to its material make-up.
  examples: solubility, fragility.

- role (http://purl.obolibrary.org/obo/BFO_0000023)
  elucidation: a role is a realizable entity that exists because of a social, institutional, or contextual circumstance.
  example: student role, patient role.

- function (http://purl.obolibrary.org/obo/BFO_0000034)
  elucidation: a function is a realizable entity that exists because it was designed (artifact) or evolved (biological) for a specific purpose.
  example: the function of a heart to pump blood, the function of a key to open a lock.

- occurrent (http://purl.obolibrary.org/obo/BFO_0000003)
  elucidation: an occurrent is an entity that unfolds in time, or a temporal/spatiotemporal region.

- process (http://purl.obolibrary.org/obo/BFO_0000015)
elucidation: a process is an occurrent that unfolds in time, has temporal parts, and can be characterized by phases, durations, inputs/outputs, start/end. It is not a tangible thing; it is an activity/event/change in which continuants participate.

- history (http://purl.obolibrary.org/obo/BFO_0000182)
- process boundary (http://purl.obolibrary.org/obo/BFO_0000035)
- spatiotemporal region (http://purl.obolibrary.org/obo/BFO_0000011)
- temporal region (http://purl.obolibrary.org/obo/BFO_0000008)
- one-dimensional temporal region (http://purl.obolibrary.org/obo/BFO_0000038)
- zero-dimensional temporal region (http://purl.obolibrary.org/obo/BFO_0000148)

---

Here is a sketch of the taxonomy of these categories:

- entity (http://purl.obolibrary.org/obo/BFO_0000001)
  - continuant (http://purl.obolibrary.org/obo/BFO_0000002)
    - generically dependent continuant (http://purl.obolibrary.org/obo/BFO_0000031)
    - independent continuant (http://purl.obolibrary.org/obo/BFO_0000004)
      - immaterial entity (http://purl.obolibrary.org/obo/BFO_0000141)
        - continuant fiat boundary (http://purl.obolibrary.org/obo/BFO_0000140)
          - fiat line (http://purl.obolibrary.org/obo/BFO_0000142)
          - fiat surface (http://purl.obolibrary.org/obo/BFO_0000146)
          - fiat point (http://purl.obolibrary.org/obo/BFO_0000147)
        - site (http://purl.obolibrary.org/obo/BFO_0000029)
        - spatial region (http://purl.obolibrary.org/obo/BFO_0000006)
          - one-dimensional spatial region (http://purl.obolibrary.org/obo/BFO_0000026)
          - two-dimensional spatial region (http://purl.obolibrary.org/obo/BFO_0000009)
          - three-dimensional spatial region (http://purl.obolibrary.org/obo/BFO_0000028)
          - zero-dimensional spatial region (http://purl.obolibrary.org/obo/BFO_0000018)
      - material entity (http://purl.obolibrary.org/obo/BFO_0000040)
        - fiat object part (http://purl.obolibrary.org/obo/BFO_0000024)
        - object (http://purl.obolibrary.org/obo/BFO_0000030)
        - object aggregate (http://purl.obolibrary.org/obo/BFO_0000027)
    - specifically dependent continuant (http://purl.obolibrary.org/obo/BFO_0000020)
      - quality (http://purl.obolibrary.org/obo/BFO_0000019)
        - relational quality (http://purl.obolibrary.org/obo/BFO_0000145)
      - realizable entity (http://purl.obolibrary.org/obo/BFO_0000017)
        - disposition (http://purl.obolibrary.org/obo/BFO_0000016)
        - function (http://purl.obolibrary.org/obo/BFO_0000034)
        - role (http://purl.obolibrary.org/obo/BFO_0000023)
        - (generic realizable entities that are not functions, roles, or dispositions remain directly under realizable entity)
  - occurrent (http://purl.obolibrary.org/obo/BFO_0000003)
    - process (http://purl.obolibrary.org/obo/BFO_0000015)
      - history (http://purl.obolibrary.org/obo/BFO_0000182)
    - process boundary (http://purl.obolibrary.org/obo/BFO_0000035)
    - spatiotemporal region (http://purl.obolibrary.org/obo/BFO_0000011)
    - temporal region (http://purl.obolibrary.org/obo/BFO_0000008)
      - one-dimensional temporal region (http://purl.obolibrary.org/obo/BFO_0000038)
      - zero-dimensional temporal region (http://purl.obolibrary.org/obo/BFO_0000148)

---

Here is bfo decision logic:

1. does the entity persist in time or unfold in time?
- persist in time → continuant
- unfold in time → occurrent

if continuant:
 2. is it a property of another entity or dependent on at least one other entity?
 - no → independent continuant
   3. does it have some portion of matter as a continuant part?
   - never → immaterial entity
     4. is its location relative to some material entity?
     - yes → how many dimensions?
       - 0 → zero-dimensional continuant boundary
       - 1 → one-dimensional continuant fiat boundary
       - 2 → two-dimensional continuant fiat boundary
       - 3 → site
     - no → is it a spatial projection of space-time?
       - yes → how many dimensions?
         - 0 → zero-dimensional spatial region
         - 1 → one-dimensional spatial region
         - 2 → two-dimensional spatial region
         - 3 → three-dimensional spatial region
   - always → material entity
     5. is it a collection of disjoint, self-standing objects?
       - yes → object aggregate
       - no → is it a proper part of an object?
         - yes → fiat object part
         - no → object
 - yes → dependent continuant
   3. can it be copied across multiple bearers?
     - yes → generically dependent continuant
     - no → specifically dependent continuant
       4. is it continuously present and directly observable or definable (including setup parameters like rate, angle, temperature)?
         - yes → quality (or relational quality if depends on ≥2 entities)
         - no → realizable entity
           - exists in a specific situation → role
           - exists by design/evolution → function
           - ceases if material make-up changes → disposition
           - otherwise → realizable entity (generic, e.g. stress, modulus of elasticity, elongation)
if occurrent:
 2. what is its relation to the universal spatiotemporal region?
 - inhabits →
   3. does it have proper temporal parts?
     - yes → process
       4. is it the sum of all processes in an entity's region?
         - yes → history
         - no → is it a collection of disjoint processes?
           - yes → process profile
     - no → process boundary
 - portion of → spatiotemporal region
 - projection of → temporal region
   3. how many dimensions?
     - 0 → zero-dimensional temporal region
     - 1 → one-dimensional temporal region

---

Operational tests for unseen terms:

1. Countability & individuation test
   - If instances are countable (“one X, two Xs”) and can be co-located with other things at an instant  → continuant → material entity → object.
   - If instances are primarily described by duration/phase (“lasting for t seconds”, “consists of steps”) → occurrent → process.

2. Modification test
   - If the term accepts “the process of …” without semantic change → likely process.
   - If the term accepts “the device/container/instrument …” without semantic change → likely object.

3. Temporal-part test
   - If parts are temporal (stages, phases) → process.
   - If parts are spatial (components, subparts) and the whole is wholly present at any time → object.

4. Syntactic signal (weak prior)
   - Gerund/verb-derived heads (…-ing; “measurement”, “operation”) bias to process.
   - Noun heads typical of apparatus (e.g., device, instrument, container, chamber, bath, vessel, tank, fixture, 
     tool, probe, sensor, holder, stand, rig) bias to object.

5. Quality vs. Object test
   - If the term denotes a state, grade, condition, or observable property of something (often adjectives like 
     hard, soft, medium, raw, cooked, brittle, elastic, flexible, transparent) → specifically dependent continuant → quality.
   - If the term denotes a thing with parts, boundaries, and independent existence (can be counted, bought, stored) → independent continuant → object.
   - Rule of thumb: If you can say “X is [term]” → quality. If you can say “an [term]” as a standalone count noun → object.

(Signals are priors; final classification must follow BFO definitions and tests above.) """.strip()

        # ----------------
        # PROMPT
        # ----------------
        def _build_hierarchy_edges(terms_with_defs: List[Dict[str, str]], parent_map: Dict[str, str]) -> List[Dict[str, str]]:
            num_to_term = {t["num"]: t for t in terms_with_defs}
            edges: List[Dict[str, str]] = []
            for child_num, parent_num in (parent_map or {}).items():
                if not parent_num:
                    continue
                if child_num not in num_to_term or parent_num not in num_to_term:
                    continue
                edges.append({
                    "child_num": child_num,
                    "child_label": num_to_term[child_num]["label"],
                    "parent_num": parent_num,
                    "parent_label": num_to_term[parent_num]["label"],
                })
            return edges

        def build_prompt(
            terms_with_defs: List[Dict[str, str]],
            parent_map: Dict[str, str],
            hierarchy_edges: List[Dict[str, str]],
            meta: Dict[str, str],
        ) -> str:
            return f"""You are an ontology engineer. Output ONLY valid Turtle (no markdown).

INPUT
- Source: Input file (classes & definitions).

- Ontology IRI: {meta['ontology_iri']}
- Entity IRI policy: use the ontology IRI with fragment '#<entitylabel>', where <entitylabel> is the lowercase class label with spaces removed.
  Example: {meta['ontology_iri']}#fracture
- Numbering hierarchy is authoritative for domain subclassing (if any).

TERMS (number, label, definition):
{json.dumps(terms_with_defs, ensure_ascii=False, indent=2)}

NUMBERING PARENT MAP (child_num → parent_num; empty string = no parent):
{json.dumps(parent_map, ensure_ascii=False, indent=2)}

LOCAL CLASS HIERARCHY (child → parent; derived from input, local classes only):
{json.dumps(hierarchy_edges, ensure_ascii=False, indent=2)}

TASKS
A) Create one domain class per term (plus optional minted intermediates if needed).
   - Labels MUST be lowercase words with spaces only (no hyphen/underscore).
   - IRI local name for each class = label with spaces removed (lowercase), appended after '#'.
     Example: label "upper yield strength" → IRI <{meta['ontology_iri']}#upperyieldstrength>.

   - For extracted terms (in the TERMS list), use the exact provided definition text as `skos:definition` (do not paraphrase). This is the "original" definition.
   - For minted classes (not in TERMS), generate a new intensional definition (1–3 sentences). This is the "original" definition.

   - For every domain class (including minted classes), generate an Aristotelian genus–differentia textual definition.
     The Aristotelian definition MUST:
       be a single English sentence,
       start with an indefinite article + the label (e.g. "a corrosion test is a process that …"),
       have the form "a/an <label> is a <genus> that <differentia>",
       use as genus the immediate domain parent class if it exists; otherwise use the selected BFO parent as genus,
       state only necessary characteristics that distinguish the class from its siblings,
       remain faithful to the meaning of the original definition.

   - In the Turtle:
       - Each class MUST have exactly two `skos:definition` literals.
       - The first `skos:definition` literal MUST contain the original definition and MUST start with "Original: ".
       - The second `skos:definition` literal MUST contain only the Aristotelian genus–differentia definition and MUST start with "Aristotelian: ".
   - For minted classes (not in TERMS), generate:
       - a new intensional source original definition (1–3 sentences), emitted as the first `skos:definition` literal starting with "Original: ",
       - an Aristotelian definition built from that original definition, emitted as the second `skos:definition` literal starting with "Aristotelian: ".

   - If multiple terms denote specific states or grades of the same property (e.g., hard/medium/soft or small/medium/big), 
      then mint a new superclass for that property (e.g., <domain> state or <domain> size). 
      Place the state terms as subclasses of this new class. 
      The minted superclass should be aligned to BFO:quality, since it represents a condition/property of the object, 
      not a material object.
   - If a definition contains explicit categorical boundaries (e.g., "- Small: < x - Medium: y - Big: > z" or "- Short: < x - Medium: y - High: > z"), 
      then mint one new superclass (e.g., "<domain> size") and create subclasses for each category (e.g., "small <domain>", "medium <domain>", "big <domain>"). 

B) Domain subclassing:
   - If a term has a parent number, add rdfs:subClassOf to the **domain class** minted for that parent term (in addition to BFO mapping).
   - If a local hierarchy edge is provided for a class, you MUST implement it as the class’s rdfs:subClassOf (domain parent). Do not contradict or drop these edges.

C) BFO mapping:
   - Add **exactly one** rdfs:subClassOf to the most specific BFO class (BFO 2020 IRIs only).
   - Use the BFO definitions, taxonomy, and decision logic provided below.
   - For each domain class (including new ones), choose the MOST SPECIFIC BFO category and use its exact IRI as rdfs:subClassOf.
   - Do NOT invent new BFO IRIs; only use those in the provided BFO text.

D) Ontology header (MUST include, values already resolved from the input and today’s date):
@prefix declarations (rdf, rdfs, owl, xsd, skos, dc, dct) ;
<{meta['ontology_iri']}> a owl:Ontology ;
    rdfs:label "{meta['label']}"@en ;
    dc:contributor "LLM" ;
    dc:license <https://creativecommons.org/publicdomain/zero/1.0/> ;
    dcterms:description "{meta['description']}"@en ;
    dcterms:title "{meta['title']}"@en ;
    owl:versionInfo "{meta['version']}" ;
    owl:imports <http://purl.obolibrary.org/obo/bfo/2020/bfo.owl> .

E) Class blocks (NO rdfs:isDefinedBy):
<CLASS_IRI> a owl:Class ;
    rdfs:label "<lowercase label>"@en ;
    skos:definition "Original: <original definition text>. " @en ;
    skos:definition "Aristotelian: <Aristotelian genus–differentia definition in the form 'a/an <label> is a <genus> that <differentia>'>."@en ;
    rdfs:subClassOf <DOMAIN_PARENT_IRI> .

STRICT RULES
- TTL only; no commentary.
- Each class MUST have exactly one rdfs:subClassOf.
- Domain parent takes precedence over BFO parent.
- If no domain parent, then use BFO.

BFO reference (2020 subset + decision logic):
{BFO_TEXT}
""".strip()

        hierarchy_edges = _build_hierarchy_edges(terms, parent_map)
        prompt = build_prompt(terms, parent_map, hierarchy_edges, META)
        system_text = "You are a BFO-aligned ontology engineer who outputs only valid Turtle syntax."

        log("Calling model...")
        raw_text = _call_llm(
            provider=provider,
            api_key=api_key,
            model=model_label,
            system_text=system_text,
            user_prompt=prompt,
        )

        ttl_output = _strip_code_fences(raw_text)

        # ===== normalize standard prefixes =====
        ttl_output = _normalize_standard_prefixes(ttl_output)

        if not ttl_output.strip().startswith("@prefix"):
            raise ValueError("Model did not return TTL starting with @prefix. Please re-run or reduce input size.")

        with open(ttl_out, "w", encoding="utf-8") as f:
            f.write(ttl_output)
        log(f"TTL saved to: {ttl_out}")

        # ===== semantic validation for the OWL import triple =====
        _validate_bfo_import(ttl_out, META["ontology_iri"])
        log("BFO import triple validated successfully (owl:imports is correct).")

        try:
            g2 = Graph()
            g2.parse(ttl_out, format="turtle")
            log("TTL parsed successfully. Ontology is syntactically valid.")
        except Exception as e:
            log(f"TTL parsing failed: {e}")

    except Exception as e:
        log(f"[Error] {e}")

    return "\n".join(logs)
